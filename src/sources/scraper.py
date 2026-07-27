#!/usr/bin/env python3
"""
Timeframe benchmark — unified scraper.

Drives data ingestion entirely from sources.yaml. No per-source bespoke files.

Per-source behaviour:

    fetch -> parse -> dedupe -> append to data/<id>/<YYYY-MM-DD>.parquet

The parser is generic: each entry's `endpoint.type` selects a fetcher (rest_json,
rest_csv, rss). The `schema` block names where the timestamp + value(s) live
inside the payload. For payloads with shapes that don't decompose neatly, the
scraper falls back to writing the raw payload as a single timestamped row
(timestamp = now, value = json string) — this is the fallback for
"snapshot" sources whose meaning is "what the API said at poll time".

Idempotency: deduplication is on the timestamp column, scoped per UTC date
file. Re-running the same minute does not duplicate.

Failures are logged to data/<id>/_errors.log with HTTP code + message.

Usage:
    python scraper.py --id <source_id>            # single source
    python scraper.py --all                       # iterate every source
    python scraper.py --domain energy             # subset
    python scraper.py --id <id> --dry-run         # fetch + parse but do not write
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import io
import json
import logging
import os
import re
import sys
import time
import zipfile
from pathlib import Path
from typing import Any, Iterable, Optional

import yaml

try:
    import httpx
except ImportError:                                # noqa: BLE001
    sys.stderr.write("scraper.py requires httpx. Install: pip install httpx pyarrow pyyaml\n")
    raise

try:
    import pyarrow as pa                           # noqa: F401
    import pyarrow.parquet as pq                   # noqa: F401
except ImportError:
    sys.stderr.write("scraper.py requires pyarrow. Install: pip install pyarrow\n")
    raise


ROOT = Path(__file__).resolve().parent
SOURCES_YAML = ROOT / "sources.yaml"
DATA_DIR = ROOT / "data"

logging.basicConfig(
    level=os.environ.get("SCRAPER_LOG_LEVEL", "INFO"),
    format="%(asctime)s %(levelname)s %(message)s",
)
log = logging.getLogger("timeframe-scraper")

UTC = dt.timezone.utc


# ──────────────────────────────────────────────────────────────────────────
# Source loading
# ──────────────────────────────────────────────────────────────────────────


def load_sources() -> list[dict]:
    with open(SOURCES_YAML) as f:
        sources = yaml.safe_load(f)
    if not isinstance(sources, list):
        raise ValueError("sources.yaml must be a YAML list")
    return sources


def get_source(sid: str) -> dict:
    for s in load_sources():
        if s["id"] == sid:
            return s
    raise KeyError(f"source id not found: {sid}")


# ──────────────────────────────────────────────────────────────────────────
# URL templating
# ──────────────────────────────────────────────────────────────────────────

_TODAY_PATTERNS = [
    ("{YYYYMMDD}", "%Y%m%d"),
    ("{YYYY-MM-DD}", "%Y-%m-%d"),
    ("{YYYY/MM/DD}", "%Y/%m/%d"),
    ("{YYYY}", "%Y"),
    ("{MM}", "%m"),
    ("{DD}", "%d"),
    ("{YYYYMM}", "%Y%m"),
]

_OFFSET_DATE_RE = re.compile(r"\{YYYY-MM-DD-(\d+)d\}")
_OFFSET_COMPACT_RE = re.compile(r"\{YYYYMMDD-(\d+)d\}")
_OFFSET_MONTH_RE = re.compile(r"\{YYYYMM-(\d+)m\}")
_OFFSET_HOUR_RE = re.compile(r"\{H-(\d+)\}")
_OFFSET_EPOCH_RE = re.compile(r"\{EPOCH-(\d+)h\}")


def _month_offset(now: dt.datetime, n: int) -> str:
    """`now` minus n calendar months, formatted YYYYMM."""
    total = now.year * 12 + (now.month - 1) - n
    return f"{total // 12:04d}{total % 12 + 1:02d}"


def expand_url(url: str, now: Optional[dt.datetime] = None) -> str:
    """Substitute date tokens in URL templates. Tokens left untouched if value
    not derivable from current UTC date.

    Static tokens: {YYYY}, {MM}, {DD}, {YYYYMMDD}, {YYYY-MM-DD},
    {YYYY/MM/DD}, {YYYYMM}, {H} (current UTC hour, 0-23, unpadded),
    {HH} (zero-padded).

    Relative-offset tokens (regex):
      {YYYY-MM-DD-Nd}   today minus N days, formatted YYYY-MM-DD
                        (e.g. {YYYY-MM-DD-30d} = 30 days ago).
      {YYYYMMDD-Nd}     same, compact — for hosts whose daily files are named
                        YYYYMMDD and whose local trading date lags UTC
                        (e.g. NYISO at 01:00 UTC is still on yesterday's file).
      {YYYYMM-Nm}       current month minus N calendar months, formatted YYYYMM.
      {H-N}             current UTC hour minus N (wraps at midnight UTC)
                        (e.g. {H-1} = previous hour).
      {EPOCH} / {EPOCH-Nh}  current unix seconds / N hours ago — for APIs whose
                        windows are epoch query params (OpenSky ?begin=&end=).

    Tokens like {ARTICLE} or {PACKAGE} or {CRATE} are *not* expanded — these
    are panel-iteration placeholders; the caller should pre-expand them.
    """
    now = now or dt.datetime.now(UTC)
    # Relative offsets first so they don't get partially consumed by the literal
    # {H} / {YYYY-MM-DD} substitutions below.
    url = _OFFSET_DATE_RE.sub(
        lambda m: (now - dt.timedelta(days=int(m.group(1)))).strftime("%Y-%m-%d"),
        url,
    )
    url = _OFFSET_COMPACT_RE.sub(
        lambda m: (now - dt.timedelta(days=int(m.group(1)))).strftime("%Y%m%d"),
        url,
    )
    url = _OFFSET_MONTH_RE.sub(lambda m: _month_offset(now, int(m.group(1))), url)
    url = _OFFSET_HOUR_RE.sub(
        lambda m: str((now - dt.timedelta(hours=int(m.group(1)))).hour),
        url,
    )
    url = _OFFSET_EPOCH_RE.sub(
        lambda m: str(int(now.timestamp()) - 3600 * int(m.group(1))),
        url,
    )
    if "{EPOCH}" in url:
        url = url.replace("{EPOCH}", str(int(now.timestamp())))
    if "{HH}" in url:
        url = url.replace("{HH}", f"{now.hour:02d}")
    if "{H}" in url:
        url = url.replace("{H}", str(now.hour))
    if "{ISO_DATE}" in url:
        url = url.replace("{ISO_DATE}", now.strftime("%Y-%m-%d"))
    if "{ISO_DATETIME}" in url:
        url = url.replace("{ISO_DATETIME}", now.strftime("%Y-%m-%dT%H:%M:%S"))
    for token, fmt in _TODAY_PATTERNS:
        if token in url:
            url = url.replace(token, now.strftime(fmt))
    return url


# ──────────────────────────────────────────────────────────────────────────
# Fetchers
# ──────────────────────────────────────────────────────────────────────────


def _resolve_auth(auth_field: Optional[str]) -> dict[str, str]:
    """Return http headers derived from an `auth:` field.

    Supported forms:
      auth: none                           → {}
      auth: env:VAR                        → header inferred from VAR suffix
                                             (*_USER_AGENT → User-Agent,
                                              *_API_KEY    → X-API-Key,
                                              *_TOKEN      → Authorization: Bearer,
                                              default      → Authorization: Bearer)
      auth: env:VAR:HeaderName             → explicit header name
                                             (e.g. env:LTA_ACCOUNT_KEY:AccountKey
                                              → AccountKey: <value>)
    """
    if not auth_field or auth_field == "none":
        return {}
    if auth_field.startswith("env:"):
        rest = auth_field[4:]
        # env:VAR:HeaderName — explicit header name
        if ":" in rest:
            var, header_name = rest.split(":", 1)
            val = os.environ.get(var)
            if not val:
                log.warning("env var %s referenced but unset; sending no auth", var)
                return {}
            return {header_name: val}
        var = rest
        val = os.environ.get(var)
        if not val:
            log.warning("env var %s referenced but unset; sending no auth", var)
            return {}
        # Inferred header from VAR suffix
        if var.endswith("USER_AGENT"):
            return {"User-Agent": val}
        if var.endswith("API_KEY"):
            return {"X-API-Key": val}
        if var.endswith("TOKEN"):
            return {"Authorization": f"Bearer {val}"}
        return {"Authorization": f"Bearer {val}"}
    return {}


def _expand_env(url: str) -> str:
    """Substitute {ENVVAR} placeholders in URL from os.environ.

    Used for APIs that require the key in the URL query string rather than a
    header (EIA `?api_key=...`, Banxico `?token=...`, USDA NASS `?key=...`).
    Unset variables are left in place so the leftover-placeholder check downstream
    catches the misconfiguration.
    """
    return _ENV_PLACEHOLDER_RE.sub(
        lambda m: os.environ.get(m.group(1), m.group(0)),
        url,
    )


_ENV_PLACEHOLDER_RE = re.compile(r"\{([A-Z][A-Z0-9_]+)\}")


# The scrape host has 4 GB of RAM and runs every source in one process, so a
# single unbounded response can OOM the box (and has). Stream every body and
# refuse anything past the cap rather than materialising it.
MAX_RESPONSE_BYTES = 48 * 1024 * 1024

# Headers describing the *encoded* transfer must not be carried onto the
# reconstructed response — the body we attach is already decoded.
_TRANSFER_HEADERS = {"content-encoding", "content-length", "transfer-encoding"}


def _capped_response(resp: httpx.Response, url: str,
                     max_bytes: Optional[int] = None) -> httpx.Response:
    """Read a streaming response into memory, refusing oversized bodies.

    Raises RuntimeError past the cap instead of truncating: a half-read JSON or
    CSV payload would parse into silently wrong records, which is worse than a
    logged failure. Raise the ceiling per source with `endpoint.max_bytes`.
    """
    cap = max_bytes or MAX_RESPONSE_BYTES
    chunks, total = [], 0
    for chunk in resp.iter_bytes():
        total += len(chunk)
        if total > cap:
            raise RuntimeError(
                f"response exceeded {cap} byte cap (still streaming at {total}): {url} "
                f"— narrow the query, or set endpoint.max_bytes if this size is expected"
            )
        chunks.append(chunk)
    hdrs = [(k, v) for k, v in resp.headers.multi_items()
            if k.lower() not in _TRANSFER_HEADERS]
    return httpx.Response(resp.status_code, headers=hdrs,
                          content=b"".join(chunks), request=resp.request)


def _http_get(url: str, headers: dict[str, str],
              max_bytes: Optional[int] = None,
              read_timeout: Optional[float] = None) -> httpx.Response:
    # Default UA acceptable for most APIs; crates.io/Reddit need a contact.
    headers = {"User-Agent": "timeframe-scraper/0.1 (+https://github.com/timeframe-bench)", **headers}
    timeout = httpx.Timeout(read_timeout or 30.0, connect=10.0)
    with httpx.Client(timeout=timeout, follow_redirects=True) as client:
        with client.stream("GET", url, headers=headers) as resp:
            return _capped_response(resp, url, max_bytes)


def _http_post(url: str, headers: dict[str, str], body: Any,
               max_bytes: Optional[int] = None,
               read_timeout: Optional[float] = None,
               form: bool = False) -> httpx.Response:
    """POST a JSON body — used by GraphQL and POST-only JSON APIs. `body` is the
    request payload (e.g. {"query": "...", "variables": {...}} for GraphQL).
    With `form=True` the body is sent application/x-www-form-urlencoded instead
    (legacy gov/utility endpoints that reject JSON payloads)."""
    headers = {"User-Agent": "timeframe-scraper/0.1 (+https://github.com/timeframe-bench)", **headers}
    timeout = httpx.Timeout(read_timeout or 30.0, connect=10.0)
    kwargs: dict[str, Any] = {"data": body} if form else {"json": body}
    with httpx.Client(timeout=timeout, follow_redirects=True) as client:
        with client.stream("POST", url, headers=headers, **kwargs) as resp:
            return _capped_response(resp, url, max_bytes)


def _paginate_json(first: bytes, url: str, headers: dict[str, str], pg: dict,
                   max_bytes: Optional[int] = None) -> bytes:
    """Follow cursor/next-link pagination for a JSON API and merge item arrays.

    `pg` config: {items: <top-level array key>, next: <top-level next-URL key>,
    max_pages: N}. Handles APIs (hubeau, OONI, ...) that return HTTP 206 + a
    `next` URL with the remaining rows. Returns one merged JSON payload."""
    items_key = pg.get("items")
    next_key = pg.get("next")
    max_pages = int(pg.get("max_pages", 5))
    try:
        doc = json.loads(first)
    except Exception:  # noqa: BLE001
        return first
    if not isinstance(doc, dict) or items_key not in doc:
        return first
    merged = list(doc.get(items_key) or [])
    nxt = doc.get(next_key) if next_key else None
    pages = 1
    while nxt and isinstance(nxt, str) and nxt.startswith("http") and pages < max_pages:
        try:
            r = _http_get(nxt, headers, max_bytes)
            if r.status_code not in (200, 206):
                break
            d = json.loads(r.content)
        except Exception:  # noqa: BLE001
            break
        merged += list(d.get(items_key) or [])
        nxt = d.get(next_key) if next_key else None
        pages += 1
    doc[items_key] = merged
    return json.dumps(doc).encode()


def _expand_panel(url: str, panel_row: dict[str, str]) -> str:
    """Substitute panel placeholders ({ARTICLE}, {PACKAGE}, {CRATE}, ...)."""
    for key, val in panel_row.items():
        url = url.replace("{" + key + "}", val)
    return url


_PLACEHOLDER_RE = re.compile(r"\{([A-Z_]+)\}")


def _expand_body(obj: Any, panel_row: Optional[dict[str, str]] = None) -> Any:
    """Recursively expand date/epoch/panel/env tokens in POST-body strings."""
    if isinstance(obj, str):
        out = expand_url(obj)
        if panel_row:
            out = _expand_panel(out, panel_row)
        return _expand_env(out)
    if isinstance(obj, dict):
        return {k: _expand_body(v, panel_row) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_expand_body(v, panel_row) for v in obj]
    return obj


def fetch_payload(src: dict, panel_row: Optional[dict[str, str]] = None) -> tuple[bytes, str]:
    ep = src["endpoint"]
    url = expand_url(ep["url"])
    if panel_row:
        url = _expand_panel(url, panel_row)
    url = _expand_env(url)                                    # {ENVVAR} → os.environ[VAR]
    leftover = _PLACEHOLDER_RE.findall(url)
    if leftover:
        raise RuntimeError(
            f"unresolved placeholders {leftover} in URL — define `panel` in sources.yaml or pass --panel"
        )
    headers = _resolve_auth(ep.get("auth"))
    # Static per-source headers (Origin/Referer/contact User-Agent — several
    # public APIs 403 without them). {ENVVAR} in values resolves from the
    # environment so contact addresses / keys stay out of the committed catalog.
    for name, val in (ep.get("headers") or {}).items():
        headers[name] = _expand_env(str(val))
    cap = ep.get("max_bytes")
    # Some hosts are simply slow (WHO GHO serves multi-MB indicator
    # payloads from a cold cache); let a source raise its own ceiling.
    rto = ep.get("timeout")
    # POST/GraphQL: type: graphql (body defaults to {"query": ...}) or an explicit
    # method: POST with a `body` JSON payload; `body_form` sends it urlencoded
    # instead. Everything else is a GET.
    form_body = ep.get("body_form")
    is_post = (ep.get("method", "").upper() == "POST") or ep.get("type") == "graphql" \
        or form_body is not None
    body = form_body if form_body is not None else ep.get("body")
    if body is None and ep.get("query"):
        body = {"query": ep["query"], "variables": ep.get("variables", {})}
    # Date/epoch tokens and panel placeholders resolve inside POST bodies too —
    # GraphQL APIs (Norway trafikkdata) take their time window in the body, not
    # the URL, and a fixed window would freeze the source on historical data.
    if body is not None:
        body = _expand_body(body, panel_row)
    # Up to 3 attempts. Two failure modes are retried: transient transport
    # faults (DNS blips, SSL handshake stalls — observed on WSL2 / long-haul
    # routes), and 429/503 throttling with Retry-After honoured (capped) —
    # pypistats in particular 429s bursts of panel rows.
    for attempt in range(3):
        try:
            resp = (_http_post(url, headers, body, cap, rto,
                               form=form_body is not None) if is_post
                    else _http_get(url, headers, cap, rto))
        except httpx.TransportError as e:
            if attempt == 2:
                raise
            wait = 5 * (attempt + 1)
            log.info("transport error for %s (%s) — retrying in %ds", url[:100], e, wait)
            time.sleep(wait)
            continue
        if resp.status_code in (429, 503) and attempt < 2:
            try:
                wait = min(float(resp.headers.get("Retry-After") or 15), 60.0)
            except ValueError:  # HTTP-date form; just use the default
                wait = 15.0
            log.info("HTTP %d for %s — retrying in %.0fs", resp.status_code, url[:100], wait)
            time.sleep(wait)
            continue
        break
    # Accept 206 Partial Content: paginated JSON APIs (hubeau, OONI) return it.
    if resp.status_code not in (200, 206):
        raise RuntimeError(f"HTTP {resp.status_code}: {resp.text[:200]}")
    content = resp.content
    if ep.get("paginate"):
        content = _paginate_json(content, url, headers, ep["paginate"], cap)
    return content, resp.headers.get("content-type", "")


# ──────────────────────────────────────────────────────────────────────────
# Parsers
# ──────────────────────────────────────────────────────────────────────────


def _tokenize(path: str) -> list[tuple[str, str]]:
    """Tokenize path syntax into [('key',name) | ('idx', '' or 'N')] pairs.

    Examples:
        'data.foo'      -> [('key','data'), ('key','foo')]
        'a[].b'         -> [('key','a'), ('idx',''), ('key','b')]
        'prices[][0]'   -> [('key','prices'), ('idx',''), ('idx','0')]
        'a[3].b'        -> [('key','a'), ('idx','3'), ('key','b')]
    """
    tokens: list[tuple[str, str]] = []
    i, n = 0, len(path)
    while i < n:
        c = path[i]
        if c == ".":
            i += 1
        elif c == "[":
            j = path.index("]", i)
            tokens.append(("idx", path[i + 1:j]))
            i = j + 1
        else:
            j = i
            while j < n and path[j] not in ".[":
                j += 1
            tokens.append(("key", path[i:j]))
            i = j
    return tokens


def _apply(obj: Any, tokens: list[tuple[str, str]]) -> Any:
    """Apply tokens to obj. Iteration tokens (`[]`) lift the rest of the path
    over every element, returning a list aligned with the parent list."""
    if not tokens:
        return obj
    kind, val = tokens[0]
    rest = tokens[1:]

    if kind == "key":
        if obj is None:
            return None
        if isinstance(obj, dict):
            return _apply(obj.get(val), rest)
        if isinstance(obj, list):
            return [_apply(item, [(kind, val)] + rest) for item in obj]
        return None

    # idx
    if val == "":                                        # iterate
        if obj is None:
            return None
        if not isinstance(obj, list):
            return None
        out = [_apply(item, rest) for item in obj]
        # A second `[]` deeper in the path yields a list per element; flatten
        # one level so `a[].b[].c` gives one flat sequence instead of nesting
        # (which collapsed multi-level feeds like SNOTEL to a single record).
        if any(isinstance(x, list) for x in out):
            flat: list = []
            for x in out:
                flat.extend(x) if isinstance(x, list) else flat.append(x)
            return flat
        return out
    # explicit index
    try:
        idx = int(val)
    except ValueError:
        return None
    if isinstance(obj, list):
        if -len(obj) <= idx < len(obj):
            return _apply(obj[idx], rest)
        return None
    if isinstance(obj, dict):                            # numeric key fallback
        return _apply(obj.get(val), rest)
    return None


def _walk(obj: Any, path: str) -> Any:
    """Walk a JSON-like object using a dotted/bracketed path."""
    if path == "":
        return obj
    return _apply(obj, _tokenize(path))


# A number is only treated as an epoch inside these ranges. This keeps *date
# labels* — a bare year (2003), YYYYMM (202606), YYYYMMDD (20260501) or
# YYYYMMDDHH (2026050100 ≈ 2.03e9) — from being silently converted into
# 1970-era timestamps, which is what used to happen to WHO GHO year values.
_EPOCH_S_RANGE = (10**8, 2 * 10**9)          # 1973-2033 in seconds
_EPOCH_MS_RANGE = (10**12, 2 * 10**12)       # 2001-2033 in milliseconds
_DECIMAL_YEAR_RE = re.compile(r"^(\d{4})\.(\d{1,6})$")


def _looks_like_epoch(n: float) -> bool:
    return (_EPOCH_S_RANGE[0] <= n < _EPOCH_S_RANGE[1]
            or _EPOCH_MS_RANGE[0] <= n < _EPOCH_MS_RANGE[1])


def _epoch_to_iso(v: Any) -> Optional[str]:
    """Convert a number-or-string timestamp into an ISO UTC string.

    Numeric-looking values are only reinterpreted as epochs when they fall in a
    plausible epoch range (see _looks_like_epoch); otherwise they are date
    labels and pass through untouched. Without that guard, quoted epochs from
    OKX/KuCoin/Bitstamp/CityBikes stayed raw in the same column as ISO stamps
    (unparseable downstream), while integer years from WHO GHO were converted
    into 1970 timestamps.
    """
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        if s.isdigit():
            n = int(s)
            return _epoch_to_iso(n) if _looks_like_epoch(n) else s
        # Decimal years ("2026.042" — NOAA GML trends files) → ISO month start.
        m = _DECIMAL_YEAR_RE.match(s)
        if m and 1900 <= int(m.group(1)) < 2100:
            year, frac = int(m.group(1)), float("0." + m.group(2))
            return f"{year}-{min(int(frac * 12) + 1, 12):02d}-01"
        # Some feeds emit a trailing 'Z' *and* an explicit offset ("...+00:00Z").
        if s.endswith("Z") and ("+" in s[10:] or "-" in s[10:]):
            return s[:-1]
        return s
    if isinstance(v, (int, float)):
        n = float(v)
        if not _looks_like_epoch(n):
            # A date label that arrived as a number — keep it verbatim.
            return str(int(n)) if float(n).is_integer() else str(n)
        if n >= _EPOCH_MS_RANGE[0]:
            n /= 1000.0
        try:
            return dt.datetime.fromtimestamp(n, UTC).isoformat()
        except (ValueError, OverflowError, OSError):
            return None
    return str(v)


def _records_from_json(data: Any, schema: dict) -> list[dict]:
    """Build a list of {timestamp, value} records from a JSON payload.

    If `value_field` is a list, each row carries a dict of named values.

    Falls back to a single row (timestamp=now, value=raw_json) if the
    declared paths don't yield aligned arrays — useful for snapshot endpoints.
    """
    ts_path = schema.get("timestamp_field", "")
    val_path = schema.get("value_field", "")
    now_iso = dt.datetime.now(UTC).isoformat()

    # Special: timestamp_field == 'now()' -> snapshot semantics
    if ts_path == "now()":
        return [{"timestamp": now_iso, "value": json.dumps(data)[:200000]}]

    # Try to walk
    try:
        ts_seq = _walk(data, ts_path) if ts_path else None
    except Exception as e:                                  # noqa: BLE001
        log.debug("ts walk failed (%s); fallback snapshot row", e)
        return [{"timestamp": now_iso, "value": json.dumps(data)[:200000]}]

    if isinstance(val_path, list):
        try:
            val_seqs = {p.split(".")[-1].strip("[]"): _walk(data, p) for p in val_path}
        except Exception:
            return [{"timestamp": now_iso, "value": json.dumps(data)[:200000]}]
    else:
        try:
            val_seqs = {"value": _walk(data, val_path)} if val_path else {}
        except Exception:
            return [{"timestamp": now_iso, "value": json.dumps(data)[:200000]}]

    if isinstance(ts_seq, list):
        # Field-based paneling: split one payload into per-entity series by
        # tagging each row with `_panel_<key>` columns (downstream groups on
        # them). `panel_field` is a path (or list of paths) aligned by index
        # with the timestamp array — e.g. NVE returns every region's weekly
        # reservoir fill in one flat array, panel_field: ['[].omrType','[].omrnr'].
        panel_paths = schema.get("panel_field")
        if isinstance(panel_paths, str):
            panel_paths = [panel_paths]
        panel_seqs = []
        for pp in (panel_paths or []):
            try:
                panel_seqs.append(("_panel_" + pp.split(".")[-1].strip("[]"), _walk(data, pp)))
            except Exception:  # noqa: BLE001
                pass
        rows: list[dict] = []
        for i, ts in enumerate(ts_seq):
            row = {"timestamp": _epoch_to_iso(ts) or now_iso}
            for name, seq in val_seqs.items():
                if isinstance(seq, list) and i < len(seq):
                    row[name] = seq[i]
                else:
                    row[name] = seq
            for pname, pseq in panel_seqs:
                if isinstance(pseq, list) and i < len(pseq):
                    row[pname] = str(pseq[i])
            rows.append(row)
        return rows

    if ts_seq is None and val_seqs:
        # snapshot with one current value
        row = {"timestamp": now_iso}
        for name, seq in val_seqs.items():
            row[name] = seq
        return [row]

    # last resort
    return [{"timestamp": _epoch_to_iso(ts_seq) or now_iso,
             "value": json.dumps(data)[:200000]}]


def _sniff_delim(line: str) -> str | None:
    """Best-effort delimiter for one line: tab, semicolon, comma, or None
    (= whitespace).

    Semicolon beats comma when it is the more frequent separator — European
    stats offices (SNB, Destatis, Eurostat bulk) ship ';'-delimited CSV in
    which ',' is the *decimal* mark, so comma-splitting would shred the rows.
    """
    if "\t" in line:
        return "\t"
    if line.count("|") > line.count(","):
        return "|"
    if line.count(";") > line.count(","):
        return ";"
    if "," in line:
        return ","
    return None


def _split_cols(line: str, delim: str | None) -> list[str]:
    line = line.lstrip("#").strip()
    if delim is None:
        return line.split()
    # Strip the surrounding quotes these exporters wrap every field in.
    return [t.strip().strip('"') for t in line.split(delim)]


_MONTH_NAMES = {name.lower(): i for i, name in enumerate(
    ("January February March April May June July August September October "
     "November December").split(), start=1)}
_MONTH_NAMES.update({k[:3]: v for k, v in list(_MONTH_NAMES.items())})


def _compose_ts(parts: list[str], hour_of_day: bool = False,
                year_week: bool = False) -> str:
    """Assemble a timestamp split across columns into ISO-ish form.

    Handles these shapes, else joins verbatim:
      * a date string + an integer hour ('2026-01-01' + '1'): common in grid /
        gov hourly CSVs (IESO 'Date','Hour'). Hour is treated as hour-ENDING
        1..24 (the usual grid convention) -> hour-of-day 0..23, unless
        `hour_of_day` (schema.compose_hour_of_day) marks it as already 0..23
        (Melbourne pedestrian counts — hour-ending would collapse 0 and 1).
      * NDBC 'YY MM DD hh mm' / CRW 'YYYY MM DD': all-numeric parts.
      * year + month number ('2026' + '3') or English month name
        ('2026' + 'January'/'Jan', CMS enrollment) -> 'YYYY-MM'.
    """
    parts = [p.strip() for p in parts]
    # year + ISO week (schema.compose_year_week — Rutgers snow lab 'Year Week'
    # rows). Needs the explicit flag: weeks 1..12 are indistinguishable from
    # months. Resolves to the ISO week's Monday.
    if (year_week and len(parts) == 2 and parts[0].isdigit()
            and len(parts[0]) == 4 and parts[1].isdigit()):
        try:
            return dt.date.fromisocalendar(
                int(parts[0]), int(parts[1]), 1).isoformat()
        except ValueError:
            return " ".join(parts)
    # date-string + integer hour (e.g. IESO 'Date'='2026-01-01', 'Hour'='1'..'24')
    if len(parts) == 2 and re.match(r"^\d{4}[-/]\d{1,2}[-/]\d{1,2}$", parts[0]) and parts[1].isdigit():
        d = parts[0].replace("/", "-")
        h = int(parts[1])
        if hour_of_day:
            hod = h % 24
        else:
            hod = (h - 1) % 24 if 1 <= h <= 24 else h % 24   # hour-ending 1..24 -> 0..23
        return f"{d}T{hod:02d}:00:00"
    if len(parts) >= 3 and all(p.isdigit() for p in parts[:3]):
        ts = f"{parts[0]:0>4}-{parts[1]:0>2}-{parts[2]:0>2}"
        if len(parts) >= 5 and parts[3].isdigit() and parts[4].isdigit():
            ts += f"T{parts[3]:0>2}:{parts[4]:0>2}:00"
        return ts
    # year + month ('Jahr','Monat' in DWD regional averages): '2026' + '3'
    if (len(parts) == 2 and parts[0].isdigit() and len(parts[0]) == 4
            and parts[1].isdigit() and 1 <= int(parts[1]) <= 12):
        return f"{parts[0]}-{int(parts[1]):02d}"
    # year + English month name ('2026' + 'January' / 'Jan'): CMS enrollment
    if len(parts) == 2 and parts[0].isdigit() and len(parts[0]) == 4:
        mon = _MONTH_NAMES.get(parts[1].lower()) or _MONTH_NAMES.get(parts[1].lower()[:3])
        if mon:
            return f"{parts[0]}-{mon:02d}"
    return " ".join(parts)


_WIDE_DATE_RE = re.compile(r"^\d{4}(-\d{2}(-\d{2})?)?$")


def _records_from_wide_csv(text: str, schema: dict) -> list[dict]:
    """Melt a wide/pivot CSV — one row per entity, one *column per date* — into
    long (timestamp, value, _panel_*) records.

    Zillow's research CSVs, and many economic/real-estate exports, ship this
    shape: `RegionID,SizeRank,RegionName,...,2000-01-31,2000-02-29,...`.

    Config:
      schema.wide.id_fields: columns identifying the entity (become _panel_*).
    Every remaining column whose name looks like a date (YYYY, YYYY-MM or
    YYYY-MM-DD) is treated as a timestamp. Blank cells are skipped.
    """
    wide = schema.get("wide") or {}
    id_fields = wide.get("id_fields") or []
    if isinstance(id_fields, str):
        id_fields = [id_fields]

    reader = csv.DictReader(io.StringIO(text))
    rows: list[dict] = []
    for rec in reader:
        panel = {"_panel_" + f: str(rec.get(f, "")) for f in id_fields}
        for col, val in rec.items():
            if col is None or col in id_fields or not _WIDE_DATE_RE.match(col.strip()):
                continue
            if val is None or val == "":
                continue
            rows.append({"timestamp": col.strip(), "value": val, **panel})
    return rows


def _records_from_csv(text: str, schema: dict) -> list[dict]:
    if schema.get("wide"):
        return _records_from_wide_csv(text, schema)
    ts_field = schema.get("timestamp_field", "").split("/")[-1]
    val_fields = schema.get("value_field")
    if isinstance(val_fields, str):
        val_fields = [val_fields]
    ts_cols = ts_field.split() if ts_field else []

    # Many gov/science "csv" feeds are really tab- or whitespace-delimited text,
    # sometimes behind a multi-line preamble or a '#'-prefixed header (NDBC,
    # GLERL, Coral Reef Watch). Locate the header: the first early line whose
    # tokens contain every schema-declared column. Comma files with a normal
    # first-line header keep the csv.DictReader path.
    lines = [l for l in text.splitlines() if l.strip()]
    if not lines:
        return []
    # A declared field may be a '/'-separated path (shortened to its last
    # segment) OR a literal column name that itself contains '/' — e.g. NYISO's
    # "LBMP ($/MWHr)". Accept a header matching either reading.
    want_short = set(ts_cols) | {vf.split("/")[-1] for vf in (val_fields or [])}
    want_full = set(ts_cols) | set(val_fields or [])
    # An explicit schema.csv_delimiter overrides sniffing — for files where the
    # heuristic misfires (e.g. ';'-delimited rows whose cells contain commas).
    forced = schema.get("csv_delimiter")
    # schema.csv_columns names the columns explicitly for files with no usable
    # header — headerless data (Rutgers snow: bare 'year week value' rows) or a
    # header whose delimiter differs from the data rows' (NMDB NEST: whitespace
    # header over ';' rows). Lines that don't split into exactly these columns
    # (including any stray header) are skipped.
    explicit_cols = schema.get("csv_columns")
    if explicit_cols:
        recs = []
        for l in lines:
            if l.lstrip().startswith("#"):
                continue
            toks = _split_cols(l, forced or _sniff_delim(l))
            if len(toks) != len(explicit_cols) or toks == list(explicit_cols):
                continue
            recs.append(dict(zip(explicit_cols, toks)))
        return _csv_recs_to_rows(recs, schema, ts_field, ts_cols, val_fields)

    header_idx, delim, header = None, None, None
    if want_short:
        for i, l in enumerate(lines[:200]):
            d = forced or _sniff_delim(l)
            toks = _split_cols(l, d)
            if want_short <= set(toks) or want_full <= set(toks):
                header_idx, delim, header = i, d, toks
                break

    if header_idx is None or delim == "," or (forced and header_idx == 0):
        start = 0 if header_idx is None else header_idx
        reader = csv.DictReader(io.StringIO("\n".join(lines[start:])),
                                delimiter=forced or ",")
        recs = list(reader)
    else:
        recs = []
        for l in lines[header_idx + 1:]:
            if l.lstrip().startswith("#"):  # NDBC units line / trailing comments
                continue
            toks = _split_cols(l, delim)
            if len(toks) < 2:
                continue
            recs.append(dict(zip(header, toks)))

    return _csv_recs_to_rows(recs, schema, ts_field, ts_cols, val_fields)


def _csv_recs_to_rows(recs: list[dict], schema: dict, ts_field: str,
                      ts_cols: list[str], val_fields: Optional[list[str]]) -> list[dict]:
    panel_cols = schema.get("panel_field")
    if isinstance(panel_cols, str):
        panel_cols = [panel_cols]
    panel_cols = [p.split("/")[-1] for p in (panel_cols or [])]

    rows = []
    for r in recs:
        if len(ts_cols) > 1 and all(c in r for c in ts_cols):
            ts_raw = _compose_ts([r[c] for c in ts_cols],
                                 hour_of_day=bool(schema.get("compose_hour_of_day")),
                                 year_week=bool(schema.get("compose_year_week")))
        else:
            ts_raw = r.get(ts_field) or next(iter(r.values()), None)
        row = {"timestamp": _epoch_to_iso(ts_raw) or ts_raw}
        for vf in (val_fields or []):
            # Exact column name wins; fall back to the last path segment.
            key = vf if vf in r else vf.split("/")[-1]
            row[key] = r.get(key)
        if not val_fields:
            row.update({k: v for k, v in r.items() if k != ts_field and k not in ts_cols})
        for pc in panel_cols:
            if pc in r:
                row["_panel_" + pc] = str(r[pc])
        rows.append(row)
    return rows


def _records_from_zip_xml(blob: bytes, schema: dict) -> list[dict]:
    """Minimal CAISO-style ZIP-XML extractor: walk every leaf XML element and
    emit a row whenever we find both an INTERVAL_START_GMT and a VALUE sibling
    inside the same parent.  Sufficient for verification and downstream parquet
    appends; replace with a proper xmlschema parser if production-grade
    fidelity is needed."""
    import xml.etree.ElementTree as ET

    rows = []
    with zipfile.ZipFile(io.BytesIO(blob)) as z:
        for name in z.namelist():
            if not name.lower().endswith(".xml"):
                continue
            tree = ET.parse(z.open(name))
            for parent in tree.iter():
                ts = None
                val = None
                for child in parent:
                    tag = child.tag.split("}")[-1]
                    if tag in ("INTERVAL_START_GMT", "DATA_ITEM_START_TIME"):
                        ts = child.text
                    elif tag in ("VALUE",):
                        val = child.text
                if ts and val:
                    rows.append({"timestamp": ts, "value": val})
    return rows


def _records_from_rss(text: str, schema: dict) -> list[dict]:
    """RSS items -> rows of {timestamp=pubDate, value=title|category}."""
    import xml.etree.ElementTree as ET

    try:
        root = ET.fromstring(text)
    except ET.ParseError as e:
        raise RuntimeError(f"RSS parse error: {e}")

    rows = []
    for item in root.iter("item"):
        ts = item.findtext("pubDate")
        val = item.findtext("category") or item.findtext("title") or ""
        rows.append({"timestamp": _epoch_to_iso(ts) or ts, "value": val})
    if not rows:
        # Atom feeds (tsunami.gov, many gov alert streams): namespaced
        # <entry> elements with <updated>/<title>/<category term=...>.
        for entry in root.iter():
            if entry.tag.rsplit("}", 1)[-1] != "entry":
                continue
            ts = val = None
            for child in entry:
                tag = child.tag.rsplit("}", 1)[-1]
                if tag == "updated":
                    ts = child.text
                elif tag == "category" and child.get("term"):
                    val = child.get("term")
                elif tag == "title" and val is None:
                    val = child.text
            rows.append({"timestamp": _epoch_to_iso(ts) or ts, "value": val or ""})
    return rows


# ──────────────────────────────────────────────────────────────────────────
# Storage (parquet append + dedupe by timestamp)
# ──────────────────────────────────────────────────────────────────────────


def write_records(sid: str, records: list[dict], dry_run: bool = False) -> int:
    if not records:
        return 0
    today = dt.datetime.now(UTC).strftime("%Y-%m-%d")
    out_dir = DATA_DIR / sid
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{today}.parquet"

    # Coerce all values to strings so concat across heterogeneous payloads is
    # never blocked by Arrow type unification. Downstream consumers cast as
    # needed; raw fidelity is preserved.
    str_records = [{k: (None if v is None else str(v)) for k, v in r.items()} for r in records]
    new_table = pa.Table.from_pylist(str_records)

    def _all_string(t: pa.Table) -> pa.Table:
        return t.cast(pa.schema([(f.name, pa.string()) for f in t.schema]))

    new_table = _all_string(new_table)
    existing = None
    if out_path.exists():
        try:
            existing = _all_string(pq.read_table(out_path))
        except Exception as e:                              # noqa: BLE001
            # A truncated parquet (process killed mid-write before the atomic
            # write below existed) would otherwise wedge this source forever,
            # since every run re-reads today's file. Quarantine and start clean
            # rather than lose all subsequent observations too.
            corrupt = out_path.with_suffix(".parquet.corrupt")
            log.error("%s: unreadable parquet %s (%s) — quarantining to %s",
                      sid, out_path.name, e, corrupt.name)
            os.replace(out_path, corrupt)
    if existing is not None:
        try:
            combined = pa.concat_tables(
                [existing, new_table], promote_options="default"
            )
        except (TypeError, pa.lib.ArrowInvalid):
            combined = pa.concat_tables([existing, new_table])
    else:
        combined = new_table

    # Dedupe on the full row so genuinely distinct measurements at the same
    # timestamp (multiple stations/zones/events) are preserved. Re-runs of the
    # same poll fold identical rows.
    df = combined.to_pandas()
    df = df.drop_duplicates(keep="first")
    n_written = len(df)
    if dry_run:
        log.info("[dry-run] %s: would write %d unique rows to %s", sid, n_written, out_path)
        return n_written
    # Write to a sibling temp file then rename. os.replace is atomic within a
    # filesystem, so a kill mid-write (OOM, cron timeout) leaves the previous
    # good file intact instead of a truncated one. 14 day-files were lost this
    # way before this was added.
    tmp = out_path.with_suffix(f".parquet.tmp{os.getpid()}")
    try:
        pq.write_table(pa.Table.from_pandas(df, preserve_index=False), tmp)
        os.replace(tmp, out_path)
    finally:
        if tmp.exists():
            tmp.unlink(missing_ok=True)
    return n_written


def log_error(sid: str, msg: str) -> None:
    out_dir = DATA_DIR / sid
    out_dir.mkdir(parents=True, exist_ok=True)
    err = out_dir / "_errors.log"
    with err.open("a") as f:
        f.write(f"{dt.datetime.now(UTC).isoformat()} {msg}\n")


# ──────────────────────────────────────────────────────────────────────────
# Driver
# ──────────────────────────────────────────────────────────────────────────


_DECIMAL_COMMA_RE = re.compile(r"^-?\d{1,3}(\.\d{3})*,\d+$")


def parse_payload(src: dict, blob: bytes, content_type: str) -> list[dict]:
    """Parse a payload into records, then apply optional aggregation.

    `schema.aggregate: {op: count|sum, bin: PT1H}` turns an irregular event
    stream (per-event rows) into a regular per-bin series — count of events (or
    sum of the value) in each time bin, preserving `_panel_*` grouping.

    `schema.drop_null_values: true` drops records whose every value field is
    null/empty — for feeds that null-pad future periods to the end of the
    current year/day (SMARD weekly files): kept, those pad rows make the max
    written timestamp always sit in the future, blinding the freshness audit."""
    recs = _dispatch_parse(src, blob, content_type)
    schema = src.get("schema", {})
    if schema.get("decimal_comma"):
        # European decimal commas ("1.234,56" / "1,395") -> dot decimals, so
        # the numeric gate and downstream casts see real numbers (Spanish
        # MITECO / Portuguese DGEG fuel prices).
        for r in recs:
            for k, v in r.items():
                if (k != "timestamp" and not k.startswith("_panel_")
                        and isinstance(v, str) and _DECIMAL_COMMA_RE.match(v)):
                    r[k] = v.replace(".", "").replace(",", ".")
    if schema.get("drop_null_values"):
        recs = [r for r in recs
                if any(v not in (None, "")
                       for k, v in r.items()
                       if k != "timestamp" and not k.startswith("_panel_"))]
    agg = schema.get("aggregate")
    return _aggregate_records(recs, agg) if agg else recs


def _aggregate_records(records: list[dict], agg: dict) -> list[dict]:
    if not records:
        return records
    import pandas as pd

    bin_iso = agg.get("bin", "PT1H")
    op = agg.get("op", "count")
    # ISO-8601 duration -> pandas offset (PT1H->1h, PT15M->15min, P1D->1D)
    m = re.match(r"^P(?:(\d+)D)?(?:T(?:(\d+)H)?(?:(\d+)M)?(?:(\d+)S)?)?$", str(bin_iso))
    if not m:
        return records
    days, hrs, mins, secs = (int(x) if x else 0 for x in m.groups())
    freq = f"{days}D" if days else f"{hrs}h" if hrs else f"{mins}min" if mins else f"{secs}s"
    df = pd.DataFrame(records)
    if "timestamp" not in df.columns:
        return records
    df["_t"] = pd.to_datetime(df["timestamp"], errors="coerce", utc=True)
    df = df[df["_t"].notna()]
    if df.empty:
        return records
    df["_bin"] = df["_t"].dt.floor(freq)
    panel_cols = [c for c in df.columns if c.startswith("_panel_")]
    val_cols = [c for c in df.columns if c not in ("timestamp", "_t", "_bin") and not c.startswith("_panel_")]
    out = []
    for key, sub in df.groupby(["_bin"] + panel_cols, sort=True):
        key = key if isinstance(key, tuple) else (key,)
        row = {"timestamp": pd.Timestamp(key[0]).isoformat()}
        if op == "sum" and val_cols:
            row["value"] = float(pd.to_numeric(sub[val_cols[0]], errors="coerce").sum())
        else:
            row["value"] = int(len(sub))
        for pc, kv in zip(panel_cols, key[1:]):
            row[pc] = kv
        out.append(row)
    return out


def _dispatch_parse(src: dict, blob: bytes, content_type: str) -> list[dict]:
    ep_type = src["endpoint"]["type"]
    schema = src.get("schema", {})

    # gzip magic 0x1f 0x8b — used by e.g. gharchive .json.gz files. Decompress
    # transparently before dispatching.
    if blob[:2] == b"\x1f\x8b":
        import gzip
        blob = gzip.decompress(blob)

    if ep_type == "ndjson_minute_counts":
        return _records_from_ndjson_minute_counts(blob)
    if ep_type == "graphql":
        return _records_from_json(json.loads(blob), schema)
    if ep_type == "rest_json":
        data = json.loads(blob)
        # SDMX-JSON detection: response with both `dataSets` and `structure` needs
        # a per-dimension-index decoder (OECD, Eurostat, ECB, IMF). SDMX-JSON 2.0
        # (current OECD) nests them under `data` and pluralises `structures`.
        if isinstance(data, dict) and isinstance(data.get("data"), dict) \
                and "dataSets" in data["data"]:
            inner = dict(data["data"])
            structs = inner.get("structures")
            if structs and "structure" not in inner:
                inner["structure"] = structs[0] if isinstance(structs, list) else structs
            data = inner
        if isinstance(data, dict) and "dataSets" in data and "structure" in data:
            return _records_from_sdmx_json(data, schema)
        # JSON-stat 2.0 (national stats offices: SCB, SSB, Statbank, StatFin,
        # Eurostat ?format=JSON): a flat `value` array over `dimension`s.
        if isinstance(data, dict) and data.get("class") == "dataset" \
                and "value" in data and "dimension" in data:
            return _records_from_jsonstat2(data, schema)
        return _records_from_json(data, schema)
    if ep_type == "rest_csv":
        # endpoint.encoding for non-UTF-8 hosts (TEPCO ships Shift_JIS).
        enc = src["endpoint"].get("encoding", "utf-8")
        text = blob.decode(enc, errors="replace")
        # Some "csv" sources actually return zipped XML (CAISO).
        if blob[:2] == b"PK":
            return _records_from_zip_xml(blob, schema)
        return _records_from_csv(text, schema)
    if ep_type in ("rss",):
        return _records_from_rss(blob.decode("utf-8", errors="replace"), schema)
    if ep_type == "rest_xml":
        return _records_from_xml(blob, schema)
    if ep_type == "html_table":
        return _records_from_html_table(blob, schema, src["endpoint"])
    if ep_type == "rest_xlsx":
        return _records_from_xlsx(blob, schema)
    if ep_type == "s3":
        # Treated as a CSV/JSON over HTTPS
        if "json" in content_type:
            return _records_from_json(json.loads(blob), schema)
        return _records_from_csv(blob.decode("utf-8", errors="replace"), schema)
    raise NotImplementedError(f"endpoint.type {ep_type!r} not handled")


# ──────────────────────────────────────────────────────────────────────────
# New parsers: XML, HTML table, XLSX, SDMX-JSON
# ──────────────────────────────────────────────────────────────────────────


def _records_from_ndjson_minute_counts(blob: bytes) -> list[dict]:
    """Aggregate an NDJSON event stream (GH Archive hourly files) into
    per-minute event counts. Events are one JSON object per line; we extract
    `created_at` with a regex instead of json-decoding every event because the
    decompressed hourly files run to hundreds of MB."""
    from collections import Counter
    counts = Counter(
        m.group(1)
        for m in re.finditer(rb'"created_at":"(\d{4}-\d{2}-\d{2}T\d{2}:\d{2})', blob)
    )
    return [
        {"timestamp": minute.decode() + ":00Z", "value": n}
        for minute, n in sorted(counts.items())
    ]


def _records_from_xml(blob: bytes, schema: dict) -> list[dict]:
    """Parse XML → list of dicts. Schema keys give slash-separated element paths.

    Example schema for FAA NASSTATUS (returns <AIRPORT_STATUS_INFORMATION>...):
        timestamp_field: AIRPORT_STATUS_INFORMATION/Update_Time
        value_field: AIRPORT_STATUS_INFORMATION/Delay_type/ARPT/Delay/Min

    Attribute-style XML (values in attributes, one element per observation —
    SDMX-ML generic, many gov feeds) uses schema.record_path + '@attr' fields:
        record_path: obs            # iterate every <obs .../> element
        timestamp_field: '@date'    # attribute of the record element
        value_field: ['@value']     # or 'child/@attr', or a child element path
    """
    import xml.etree.ElementTree as ET
    root = ET.fromstring(blob)
    ts_path = schema.get("timestamp_field", "")
    val_paths = schema.get("value_field", [])
    if isinstance(val_paths, str):
        val_paths = [val_paths]

    record_path = schema.get("record_path")
    if record_path:
        def _field(elt, path: str) -> Optional[str]:
            if path.startswith("@"):
                return elt.get(path[1:])
            if "/@" in path:
                child_path, attr = path.rsplit("/@", 1)
                child = elt.find(child_path)
                return child.get(attr) if child is not None else None
            child = elt.find(path)
            return child.text if child is not None else None

        # Strip namespaces from tag matching: iterate by local name.
        local = record_path.split("/")[-1]
        out: list[dict] = []
        for elt in root.iter():
            if elt.tag.split("}")[-1] != local:
                continue
            rec: dict = {"timestamp": _epoch_to_iso(_field(elt, ts_path))
                         or _field(elt, ts_path)}
            for vp in val_paths:
                key = vp.lstrip("@").split("/")[-1].lstrip("@") or "value"
                rec[key] = _field(elt, vp)
            pf = schema.get("panel_field")
            for p in ([pf] if isinstance(pf, str) else (pf or [])):
                got = _field(elt, p)
                if got is not None:
                    rec["_panel_" + p.lstrip("@").split("/")[-1].lstrip("@")] = str(got)
            out.append(rec)
        return out

    def _find_all(node, path: str) -> list:
        # xpath-lite: consume the root name if it matches, then use the tail as a
        # relative path.
        parts = [p for p in path.split("/") if p]
        if parts and node.tag == parts[0]:
            parts = parts[1:]
        return list(node.iter(parts[-1])) if parts else []

    def _text(node, path: str) -> Optional[str]:
        elts = _find_all(node, path)
        return elts[0].text if elts else None

    ts_val = _text(root, ts_path) if ts_path else None
    records: list[dict] = []
    if val_paths:
        # Iterate value elements as siblings; parallel-align by index.
        value_elts = _find_all(root, val_paths[0])
        for elt in value_elts:
            rec = {"timestamp": ts_val}
            try:
                rec["value"] = float(elt.text) if elt.text else None
            except (ValueError, TypeError):
                rec["value"] = elt.text
            records.append(rec)
    if not records:
        # No value elts found → single-row snapshot at the timestamp.
        records = [{"timestamp": ts_val, "value": None}]
    return records


def _records_from_html_table(blob: bytes, schema: dict, endpoint: dict) -> list[dict]:
    """Parse an HTML page → rows from a table matching endpoint.table_selector.

    Schema:
        timestamp_field: <column name in the header row>
        value_field: <column name in the header row>
    Endpoint (optional):
        table_selector: CSS selector for the target table (default: first table)
    """
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(blob, "lxml")
    selector = endpoint.get("table_selector", "table")
    table = soup.select_one(selector)
    if table is None:
        return []
    rows = table.find_all("tr")
    if not rows:
        return []
    headers = [th.get_text(strip=True) for th in rows[0].find_all(["th", "td"])]
    ts_col = schema.get("timestamp_field", headers[0] if headers else "")
    val_col = schema.get("value_field", headers[1] if len(headers) > 1 else "")
    if isinstance(val_col, list):
        val_col = val_col[0]
    records: list[dict] = []
    for tr in rows[1:]:
        cells = [td.get_text(strip=True) for td in tr.find_all(["td", "th"])]
        if len(cells) < len(headers):
            continue
        row = dict(zip(headers, cells))
        try:
            val = float(row.get(val_col, "").replace(",", ""))
        except (ValueError, AttributeError):
            val = row.get(val_col)
        records.append({"timestamp": row.get(ts_col), "value": val})
    return records


def _records_from_xlsx(blob: bytes, schema: dict) -> list[dict]:
    """Parse .xlsx → rows using openpyxl. Reads the first sheet's first two
    columns as (timestamp, value) unless schema overrides column names.
    """
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        return []
    ts_col = schema.get("timestamp_field", header[0])
    val_col = schema.get("value_field", header[1] if len(header) > 1 else None)
    if isinstance(val_col, list):
        val_col = val_col[0]
    ts_i = header.index(ts_col) if ts_col in header else 0
    val_i = header.index(val_col) if val_col in header else 1
    records: list[dict] = []
    for row in rows_iter:
        if row is None or all(c is None for c in row):
            continue
        ts = row[ts_i] if ts_i < len(row) else None
        val = row[val_i] if val_i < len(row) else None
        try:
            val = float(val) if val is not None else None
        except (ValueError, TypeError):
            pass
        records.append({"timestamp": str(ts) if ts is not None else None, "value": val})
    return records


def _records_from_sdmx_json(data: dict, schema: dict) -> list[dict]:
    """Decode SDMX-JSON → flat rows. Used by OECD, Eurostat, ECB, IMF.

    SDMX-JSON stores observations under `dataSets[0].series` keyed by
    dimension-index strings (e.g. "0:2:1"), with each observation as
    `{"time_idx": [value, ...]}`. The time-period lookup lives in
    `structure.dimensions.observation`.
    """
    if "dataSets" not in data or "structure" not in data:
        return []
    ds = data["dataSets"][0] if data["dataSets"] else {}
    struct = data["structure"]

    # Time labels: structure.dimensions.observation is a list; the entry with
    # id == "TIME_PERIOD" gives the labels.
    obs_dims = struct.get("dimensions", {}).get("observation", [])
    time_labels: list[str] = []
    for d in obs_dims:
        if d.get("id") == "TIME_PERIOD":
            time_labels = [v.get("id", "") for v in d.get("values", [])]
            break

    records: list[dict] = []
    series = ds.get("series", {})
    if series:
        for _key, s in series.items():
            for obs_key, val in (s.get("observations") or {}).items():
                try:
                    idx = int(obs_key.split(":")[0])
                except (ValueError, IndexError):
                    continue
                t = time_labels[idx] if 0 <= idx < len(time_labels) else str(idx)
                v = val[0] if isinstance(val, list) and val else val
                records.append({"timestamp": t, "value": v})
    else:
        # Some SDMX-JSON responses put observations flat on dataSets[0].observations.
        for obs_key, val in (ds.get("observations") or {}).items():
            try:
                idx = int(obs_key.split(":")[-1])
            except (ValueError, IndexError):
                continue
            t = time_labels[idx] if 0 <= idx < len(time_labels) else str(idx)
            v = val[0] if isinstance(val, list) and val else val
            records.append({"timestamp": t, "value": v})
    return records


def _records_from_jsonstat2(data: dict, schema: dict) -> list[dict]:
    """Decode JSON-stat 2.0 (`class: dataset`) → flat rows.

    Used by national statistics offices (SCB Sweden, SSB Norway, Statbank
    Denmark, StatFin, Eurostat ?format=JSON). Layout: a flat `value` array in
    C-order over the dimensions listed in `id`, with per-dimension category
    index maps under `dimension.<id>.category.index`. The time dimension is
    named in `role.time` (falls back to a Tid/TIME/time-ish id). Non-time
    dimensions with size > 1 become `_panel_<dimid>` columns.
    """
    dims = data.get("id") or list((data.get("dimension") or {}).keys())
    sizes = data.get("size") or []
    values = data.get("value")
    dimension = data.get("dimension") or {}
    if not dims or not sizes or values is None or len(dims) != len(sizes):
        return []
    # json-stat allows value as a dict {flat_index: v} for sparse data
    if isinstance(values, dict):
        flat = [None] * (1 if not sizes else _prod(sizes))
        for k, v in values.items():
            try:
                flat[int(k)] = v
            except (ValueError, IndexError):
                pass
        values = flat

    role = data.get("role") or {}
    time_dim = (role.get("time") or [None])[0]
    if time_dim not in dims:
        time_dim = next((d for d in dims if str(d).lower() in
                         ("tid", "time", "time_period", "date", "period")), None)
    if time_dim is None:
        return []
    time_pos = dims.index(time_dim)

    def _idx_to_label(dim_id: str) -> list[str]:
        idx = ((dimension.get(dim_id) or {}).get("category") or {}).get("index") or {}
        if isinstance(idx, dict):
            return [k for k, _ in sorted(idx.items(), key=lambda kv: kv[1])]
        return list(idx)  # already an ordered list

    time_labels = _idx_to_label(time_dim)
    # panel = non-time dims with size > 1
    panel_dims = [d for i, d in enumerate(dims) if i != time_pos and sizes[i] > 1]
    panel_labels = {d: _idx_to_label(d) for d in panel_dims}

    # C-order strides
    strides = [1] * len(sizes)
    for k in range(len(sizes) - 2, -1, -1):
        strides[k] = strides[k + 1] * sizes[k + 1]

    records: list[dict] = []
    for f in range(len(values)):
        v = values[f]
        if v is None:
            continue
        ti = (f // strides[time_pos]) % sizes[time_pos]
        t = time_labels[ti] if 0 <= ti < len(time_labels) else str(ti)
        row = {"timestamp": _norm_pxweb_time(t), "value": v}
        for d in panel_dims:
            di = dims.index(d)
            ci = (f // strides[di]) % sizes[di]
            labs = panel_labels[d]
            row[f"_panel_{d}"] = labs[ci] if 0 <= ci < len(labs) else str(ci)
        records.append(row)
    return records


_DUR_RE = re.compile(
    r"^P(?:(?P<y>\d+)Y)?(?:(?P<mo>\d+)M)?(?:(?P<w>\d+)W)?(?:(?P<d>\d+)D)?"
    r"(?:T(?:(?P<h>\d+)H)?(?:(?P<mi>\d+)M)?(?:(?P<s>\d+)S)?)?$"
)
_DUR_UNITS = {"y": 31_536_000, "mo": 2_592_000, "w": 604_800, "d": 86_400,
              "h": 3600, "mi": 60, "s": 1}

# Sources at or below this period are refetched on every sweep; slower ones are
# refreshed at most this often.
_MIN_REFRESH_SECONDS = 3600
_MAX_REFRESH_SECONDS = 6 * 3600


def _period_seconds(freq: str) -> Optional[int]:
    """ISO-8601 duration -> seconds. None if unparseable (e.g. 'irregular')."""
    m = _DUR_RE.match(str(freq or "").strip())
    if not m or not any(m.groupdict().values()):
        return None
    return sum(int(v) * _DUR_UNITS[k] for k, v in m.groupdict().items() if v)


def _last_scraped_age(sid: str) -> Optional[float]:
    """Seconds since this source last wrote a parquet file; None if never."""
    d = DATA_DIR / sid
    if not d.is_dir():
        return None
    mtimes = [p.stat().st_mtime for p in d.glob("*.parquet")]
    if not mtimes:
        return None
    return time.time() - max(mtimes)


def is_due(src: dict) -> bool:
    """Should this source be fetched on the current sweep?

    The full sweep runs every 15 minutes, so anything hourly-or-faster is
    always due. Slower feeds (daily/weekly/monthly) do not gain anything from
    96 fetches a day — and each write re-reads and dedupes the whole day-file,
    so refetching a large monthly panel that often is the dominant cost of a
    sweep. Those are refreshed at most every 6 hours, which still gives several
    chances a day to pick up revisions.
    """
    period = _period_seconds(src.get("frequency"))
    if period is None or period <= _MIN_REFRESH_SECONDS:
        return True
    age = _last_scraped_age(src["id"])
    if age is None:
        return True                       # never scraped — always fetch
    return age >= min(period, _MAX_REFRESH_SECONDS)


def _norm_pxweb_time(t: str) -> str:
    """Normalise PxWeb/JSON-stat time labels to ISO-ish so downstream can parse.

    1980M01 -> 1980-01 ; 2020K2 -> 2020-Q2 ; 2020Q2 -> 2020-Q2 ;
    2020W03 -> 2020-W03 ; plain YYYY / YYYY-MM(-DD) pass through unchanged.
    """
    s = str(t)
    m = re.match(r"^(\d{4})M(\d{1,2})$", s)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}"
    m = re.match(r"^(\d{4})[KQ]([1-4])$", s)
    if m:
        return f"{m.group(1)}-Q{m.group(2)}"
    m = re.match(r"^(\d{4})W(\d{1,2})$", s)
    if m:
        return f"{m.group(1)}-W{int(m.group(2)):02d}"
    return s


def _prod(xs: list[int]) -> int:
    p = 1
    for x in xs:
        p *= int(x)
    return p


def run_one(sid: str, dry_run: bool = False) -> int:
    src = get_source(sid)
    if src.get("disabled"):
        log.info("skip %s (disabled: %s)", sid, src.get("disabled_reason", "no reason given"))
        return 0
    log.info("fetch %s -> %s", sid, src["endpoint"]["url"][:120])

    panel = src.get("panel") or [None]   # list of dicts or [None] for single-poll
    total_written = 0
    for panel_row in panel:
        try:
            blob, ct = fetch_payload(src, panel_row=panel_row)
        except Exception as e:                              # noqa: BLE001
            log.error("fetch failed for %s%s: %s", sid, f" [{panel_row}]" if panel_row else "", e)
            log_error(sid, f"FETCH {e} {panel_row or ''}")
            continue
        try:
            rows = parse_payload(src, blob, ct)
            # `panel_concat: true` marks a panel that paginates ONE series over
            # time (e.g. one CSV per calendar year) rather than enumerating
            # distinct series — skip the `_panel_` tagging so the chunks merge.
            if panel_row and not src.get("panel_concat"):
                for r in rows:
                    r.update({"_panel_" + k: v for k, v in panel_row.items()})
        except Exception as e:                              # noqa: BLE001
            log.error("parse failed for %s%s: %s", sid, f" [{panel_row}]" if panel_row else "", e)
            log_error(sid, f"PARSE {e} {panel_row or ''}")
            continue
        n = write_records(sid, rows, dry_run=dry_run)
        log.info("%s%s: wrote %d rows (raw %d)", sid, f" [{panel_row}]" if panel_row else "", n, len(rows))
        total_written += n
    return total_written


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--id", help="single source id from sources.yaml")
    ap.add_argument("--all", action="store_true", help="run every source")
    ap.add_argument("--domain", help="only run sources in this domain")
    ap.add_argument("--dry-run", action="store_true",
                    help="fetch + parse but do not write parquet")
    ap.add_argument("--list", action="store_true", help="print all source ids and exit")
    ap.add_argument("--force", action="store_true",
                    help="with --all, ignore cadence and refetch every source")
    args = ap.parse_args()

    sources = load_sources()
    if args.list:
        for s in sources:
            flag = "  [DISABLED]" if s.get("disabled") else ""
            print(f"{s['id']:35s}  {s['domain']:13s}  {s['frequency']}{flag}")
        return 0

    targets: Iterable[str]
    if args.id:
        targets = [args.id]
    elif args.all:
        # Skip slow feeds already refreshed recently — see is_due().
        due = sources if args.force else [s for s in sources if is_due(s)]
        skipped = len(sources) - len(due)
        if skipped:
            log.info("cadence: %d/%d sources due this sweep (%d not yet stale)",
                     len(due), len(sources), skipped)
        targets = [s["id"] for s in due]
    elif args.domain:
        targets = [s["id"] for s in sources if s["domain"] == args.domain]
    else:
        ap.print_help()
        return 1

    failed = 0
    for sid in targets:
        try:
            run_one(sid, dry_run=args.dry_run)
        except Exception as e:                              # noqa: BLE001
            log.exception("%s: unrecoverable error", sid)
            log_error(sid, f"UNCAUGHT {e}")
            failed += 1
    return 0 if failed == 0 else 2


if __name__ == "__main__":
    sys.exit(main())
